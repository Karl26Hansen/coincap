from openpyxl import Workbook
from request.request import get_data, display_coin_data, get_datetime_in_seconds

workbook = Workbook()
worksheet = workbook.active

data = get_data()
fifty_coins = display_coin_data(get_data())

symbols = []
names = []
coins = []
for coin in fifty_coins:
    symbols.append(coin.get_symbol())
    names.append(coin.get_name())
    coins.append(coin.get_price_usd())

worksheet.cell(1, 1, "symbol")
worksheet.cell(1, 2, "name")
worksheet.cell(1, 3, "price usd")

col = 1
for i in range(len(symbols)):
    col += 1
    cell_to_write = worksheet.cell(col, 1)
    cell_to_write.value = symbols[i]

col = 1
for i in range(len(names)):
    col += 1
    cell_to_write = worksheet.cell(col, 2)
    cell_to_write.value = names[i]

col = 1
for i in range(len(coins)):
    col += 1
    cell_to_write = worksheet.cell(col, 3)
    cell_to_write.value = coins[i]

timestamp_tuple = data.popitem()

file_name = get_datetime_in_seconds(timestamp_tuple[1] / 1000)
workbook.save(file_name + ".xlsx")
